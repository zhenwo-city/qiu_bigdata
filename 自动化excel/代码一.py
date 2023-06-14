'''
实现了跨工作簿进行整表复制，缺点是需要保存的工作表没有合并单元格
'''
from openpyxl import load_workbook
import pandas as pd

input_path = r'C:\Users\库\Desktop\input_book.xlsx'
# 读取输入的工作簿内容
work_book = load_workbook(input_path)
# 选择需要修改的工作簿sheet
work_sheet = work_book['VIP']
# 读取原始工作簿信息
data = pd.read_excel(r'C:\Users\库\Desktop\output_book.xls', header=None, sheet_name='客户 (新)')
for i in range(len(data)):
    ata_list = data.iloc[i].tolist()
    for j in range(len(ata_list)):
        # 对工作簿中的每一个单元格进行修改
        work_sheet.cell(row=i + 1, column=j + 1).value = ata_list[j]
# 保存一个新的工作簿
work_book.save(filename='./new.xlsx')
# 关闭接口
work_book.close()
